#!/usr/bin/env python3
"""Build high-resolution, Google-Slides-safe charts in Indeed brand style for the Virtue market overview deck."""
import os, glob, csv
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.font_manager as fm
from matplotlib.ticker import FuncFormatter

# ---- Indeed Sans ----
for f in glob.glob('/tmp/indeedsans/IndeedSans_*.ttf'):
    fm.fontManager.addfont(f)
plt.rcParams['font.family'] = 'Indeed Sans'
plt.rcParams['svg.fonttype'] = 'none'
plt.rcParams['axes.unicode_minus'] = False

# ---- Indeed palette ----
BLUE    = '#003A9B'   # Indeed primary blue
NAVY    = '#00123C'   # deepest navy
A11Y    = '#5592FD'   # a11y / light blue
MAGENTA = '#9D2B6B'
PURPLE  = '#5644BF'
TEAL    = '#237EA3'
ORANGE  = '#B4602B'
GREY    = '#C7CCD6'   # muted grey for non-highlighted bars
DARK    = '#2D2D2D'   # body text
LGREY   = '#E7EAF0'

OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'charts')
os.makedirs(OUT, exist_ok=True)
DPI = 220

def style_ax(ax):
    for s in ['top', 'right']:
        ax.spines[s].set_visible(False)
    for s in ['left', 'bottom']:
        ax.spines[s].set_color('#B9BFCC')
    ax.tick_params(colors=DARK, length=0, labelsize=10)
    ax.set_axisbelow(True)

def save(fig, name):
    fig.savefig(os.path.join(OUT, name), dpi=DPI, bbox_inches='tight',
                facecolor='white', pad_inches=0.12)
    plt.close(fig)
    print('saved', name)

# ======================================================================
# Load month-over-month market data
# ======================================================================
wb = openpyxl.load_workbook('Market Data (1) copy.xlsx', data_only=True)
ws = wb['Month over month Market Analysi']
hdr = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
mom = []
for r in range(2, ws.max_row + 1):
    d = {hdr[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
    if d['time (1mo)']:
        mom.append(d)

MONTHS = ['Jan','Feb','Mar','Apr','May','Jun','Jul','Aug','Sep','Oct','Nov','Dec']
def label(t):
    # t like '[2024-01-01 00:00:00, ...'
    ymd = t.split(',')[0].strip('[').strip()
    y, m, _ = ymd.split(' ')[0].split('-')
    return f"{MONTHS[int(m)-1]} {y[2:]}"

labels   = [label(d['time (1mo)']) for d in mom]
spend    = [float(d['spend_local'])/1000 for d in mom]   # €k
cpc      = [float(d['CPC']) for d in mom]
as_rate  = [float(d['Applystart_rate'])*100 for d in mom]
as_job   = [float(d['AS_per_job']) for d in mom]
inv_job  = [float(d['investment_per_job']) for d in mom]
jobs     = [float(d['Jobs']) for d in mom]
x = list(range(len(labels)))
# show every 2nd label to avoid crowding
xt = [i for i in x if i % 2 == 0]
xl = [labels[i] for i in xt]

# ---------------------------------------------------------------- Chart A
# Market spend rising + cost per click rising = intensifying competition
fig, ax1 = plt.subplots(figsize=(8.4, 4.1))
ax1.bar(x, spend, color=A11Y, width=0.72, label='Monthly market spend (€k)', zorder=2)
style_ax(ax1)
ax1.set_ylabel('Market spend (€000s)', color=DARK, fontsize=10.5)
ax1.set_xticks(xt); ax1.set_xticklabels(xl, rotation=0, fontsize=8.7)
ax1.set_ylim(0, max(spend)*1.18)
ax1.grid(axis='y', color=LGREY, linewidth=0.9)
ax2 = ax1.twinx()
ax2.plot(x, cpc, color=BLUE, linewidth=3, marker='o', markersize=4.5,
         label='Cost per click (€)', zorder=4)
ax2.spines['top'].set_visible(False)
ax2.tick_params(colors=DARK, length=0, labelsize=10)
ax2.set_ylabel('Cost per click (€)', color=BLUE, fontsize=10.5)
ax2.set_ylim(0, max(cpc)*1.35)
# annotate endpoints
ax1.annotate(f"€{spend[0]:.0f}k", (0, spend[0]), textcoords='offset points',
             xytext=(0, 5), ha='center', fontsize=8.5, color=BLUE)
ax1.annotate(f"€{spend[-1]:.0f}k", (x[-1], spend[-1]), textcoords='offset points',
             xytext=(0, 5), ha='center', fontsize=9, color=BLUE, weight='bold')
ax2.annotate(f"€{cpc[-1]:.2f}", (x[-1], cpc[-1]), textcoords='offset points',
             xytext=(0, 8), ha='center', fontsize=9, color=BLUE, weight='bold')
l1, lab1 = ax1.get_legend_handles_labels()
l2, lab2 = ax2.get_legend_handles_labels()
ax1.legend(l1 + l2, lab1 + lab2, loc='upper left', frameon=False, fontsize=9.5)
save(fig, 'A_market_spend_cpc.png')

# ---------------------------------------------------------------- Chart B
# Candidate yield falling: apply-starts per job down while cost per job up
fig, ax1 = plt.subplots(figsize=(8.4, 4.1))
ax1.plot(x, inv_job, color=ORANGE, linewidth=3, marker='o', markersize=4,
         label='Investment per job (€)')
style_ax(ax1)
ax1.set_ylabel('Investment per job (€)', color=ORANGE, fontsize=10.5)
ax1.set_xticks(xt); ax1.set_xticklabels(xl, fontsize=8.7)
ax1.set_ylim(0, max(inv_job)*1.25)
ax1.grid(axis='y', color=LGREY, linewidth=0.9)
ax2 = ax1.twinx()
ax2.plot(x, as_job, color=BLUE, linewidth=3, marker='s', markersize=4,
         label='Apply-starts per job')
ax2.spines['top'].set_visible(False)
ax2.tick_params(colors=DARK, length=0, labelsize=10)
ax2.set_ylabel('Apply-starts per job', color=BLUE, fontsize=10.5)
ax2.set_ylim(0, max(as_job)*1.3)
ax1.annotate(f"€{inv_job[0]:.0f}", (0, inv_job[0]), textcoords='offset points',
             xytext=(0, 6), ha='center', fontsize=8.5, color=ORANGE)
ax1.annotate(f"€{inv_job[-1]:.0f}", (x[-1], inv_job[-1]), textcoords='offset points',
             xytext=(-2, 6), ha='center', fontsize=9, color=ORANGE, weight='bold')
ax2.annotate(f"{as_job[0]:.0f}", (0, as_job[0]), textcoords='offset points',
             xytext=(0, -14), ha='center', fontsize=8.5, color=BLUE)
ax2.annotate(f"{as_job[-1]:.0f}", (x[-1], as_job[-1]), textcoords='offset points',
             xytext=(0, -14), ha='center', fontsize=9, color=BLUE, weight='bold')
l1, lab1 = ax1.get_legend_handles_labels()
l2, lab2 = ax2.get_legend_handles_labels()
ax1.legend(l1 + l2, lab1 + lab2, loc='upper center', frameon=False, fontsize=9.5, ncol=2)
save(fig, 'B_candidate_yield.png')

# ======================================================================
# Company-level competitive data
# ======================================================================
def load(name):
    ws = wb[name]
    h = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    rows = []
    for r in range(2, ws.max_row + 1):
        d = {h[c - 1]: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if d['companyid']:
            rows.append(d)
    return rows

def num(v):
    try: return float(v)
    except: return 0.0

cur = load('Jan to May 2026')
def shortname(c):
    return c.rsplit('[', 1)[0].strip()

# ---------------------------------------------------------------- Chart C
# Competitive landscape: top spenders Jan-May 2026, Virtue highlighted
tops = sorted(cur, key=lambda r: num(r['spend_local']), reverse=True)[:12]
virt = [r for r in cur if 'virtue' in r['companyid'].lower()][0]
rows_c = tops + [virt]
names_c = [shortname(r['companyid']) for r in rows_c]
vals_c  = [num(r['spend_local'])/1000 for r in rows_c]
order = sorted(range(len(rows_c)), key=lambda i: vals_c[i])
names_c = [names_c[i] for i in order]; vals_c = [vals_c[i] for i in order]
rows_o  = [rows_c[i] for i in order]
colors_c = [BLUE if 'virtue' in r['companyid'].lower() else GREY for r in rows_o]
fig, ax = plt.subplots(figsize=(8.6, 4.6))
bars = ax.barh(range(len(names_c)), vals_c, color=colors_c, zorder=2)
ax.set_yticks(range(len(names_c)))
ax.set_yticklabels(names_c, fontsize=9.3)
# highlight Virtue label
for t, r in zip(ax.get_yticklabels(), rows_o):
    if 'virtue' in r['companyid'].lower():
        t.set_color(BLUE); t.set_weight('bold')
style_ax(ax)
ax.set_xlabel('Spend, Jan\u2013May 2026 (€000s)', fontsize=10.5)
ax.grid(axis='x', color=LGREY, linewidth=0.9)
ax.set_xlim(0, max(vals_c)*1.12)
for i, v in enumerate(vals_c):
    ax.text(v + max(vals_c)*0.01, i, f"€{v:.0f}k", va='center', fontsize=8.7,
            color=(BLUE if colors_c[i]==BLUE else '#6A7180'),
            weight=('bold' if colors_c[i]==BLUE else 'normal'))
save(fig, 'C_competitive_spend.png')

# ---------------------------------------------------------------- Chart D
# Apply-start rate: Virtue vs market vs leading spenders
prev = load('July to December 2025')
def market_rate(rows):
    return sum(num(r['Tot_Applystarts']) for r in rows)/sum(num(r['Clicks']) for r in rows)*100
mkt_rate = market_rate(cur)
# pick a few notable competitors + virtue
pick_names = [('Comfort Keepers','Comfort\nKeepers'),('Mowlam Healthcare','Mowlam'),
              ('Dovida','Dovida'),('Be Independent','Be\nIndependent'),('Danu Home Care','Danu\nHome Care')]
comp = []
for pn, short in pick_names:
    for r in cur:
        if pn.lower() in r['companyid'].lower():
            comp.append((short, num(r['Applystart_rate'])*100)); break
labels_d = ['Virtue'] + [c[0] for c in comp] + ['Market\navg']
vals_d   = [num(virt['Applystart_rate'])*100] + [c[1] for c in comp] + [mkt_rate]
colors_d = [BLUE] + [GREY]*len(comp) + [NAVY]
fig, ax = plt.subplots(figsize=(8.6, 4.2))
b = ax.bar(range(len(labels_d)), vals_d, color=colors_d, width=0.66, zorder=2)
style_ax(ax)
ax.set_xticks(range(len(labels_d)))
ax.set_xticklabels(labels_d, fontsize=9.2)
ax.get_xticklabels()[0].set_color(BLUE); ax.get_xticklabels()[0].set_weight('bold')
ax.set_ylabel('Apply-start rate (%)', fontsize=10.5)
ax.grid(axis='y', color=LGREY, linewidth=0.9)
ax.set_ylim(0, max(vals_d)*1.2)
for i, v in enumerate(vals_d):
    ax.text(i, v + max(vals_d)*0.02, f"{v:.1f}%", ha='center', fontsize=9.4,
            color=(BLUE if colors_d[i]==BLUE else DARK),
            weight=('bold' if i==0 else 'normal'))
ax.axhline(mkt_rate, color=NAVY, linewidth=1, linestyle=(0,(4,3)), zorder=1)
save(fig, 'D_applystart_rate.png')

# ---------------------------------------------------------------- Chart E
# Virtue efficiency improvement H2 2025 -> Jan-May 2026
vp = [r for r in prev if 'virtue' in r['companyid'].lower()][0]
metrics = ['Cost per click (€)', 'Cost per\napply-start (€)', 'Apply-start rate (%)']
h2  = [num(vp['CPC']), num(vp['CPAs']), num(vp['Applystart_rate'])*100]
h1  = [num(virt['CPC']), num(virt['CPAs']), num(virt['Applystart_rate'])*100]
fig, ax = plt.subplots(figsize=(8.2, 4.2))
import numpy as np
xi = np.arange(len(metrics)); w = 0.36
ax.bar(xi - w/2, h2, w, label='H2 2025', color=GREY, zorder=2)
ax.bar(xi + w/2, h1, w, label='Jan\u2013May 2026', color=BLUE, zorder=2)
style_ax(ax)
ax.set_xticks(xi); ax.set_xticklabels(metrics, fontsize=9.6)
ax.grid(axis='y', color=LGREY, linewidth=0.9)
ax.legend(frameon=False, fontsize=10, loc='upper right')
mx = max(h2 + h1)
ax.set_ylim(0, mx*1.18)
for xx, v in zip(xi - w/2, h2):
    ax.text(xx, v + mx*0.02, f"{v:.2f}", ha='center', fontsize=8.8, color='#6A7180')
for xx, v in zip(xi + w/2, h1):
    ax.text(xx, v + mx*0.02, f"{v:.2f}", ha='center', fontsize=9, color=BLUE, weight='bold')
save(fig, 'E_virtue_efficiency.png')

# ======================================================================
# County-level candidate demand (query1 csv)
# ======================================================================
codes = {'c':'Cork','ce':'Clare','cn':'Cavan','cw':'Carlow','d':'Dublin','dl':'Donegal',
'g':'Galway','ke':'Kildare','kk':'Kilkenny','ky':'Kerry','ld':'Longford','lh':'Louth',
'lk':'Limerick','lm':'Leitrim','ls':'Laois','mh':'Meath','mn':'Monaghan','mo':'Mayo',
'oy':'Offaly','rn':'Roscommon','so':'Sligo','ta':'Tipperary','wd':'Waterford',
'wh':'Westmeath','ww':'Wicklow','wx':'Wexford'}
crows = []
with open('query1 (77).csv') as f:
    for d in csv.DictReader(f):
        if d['jladmin1code'] == 'unknown':
            continue
        d['county'] = codes.get(d['jladmin1code'], d['jladmin1code'])
        for k in ['CLICKS','JOBS','COMPANIES','JOBSEEKERS','jobseekers_per_job']:
            d[k] = float(d[k])
        crows.append(d)

# ---------------------------------------------------------------- Chart F
# ALL 26 counties by candidate volume (jobseekers); colour = supply tightness.
# Two equal columns (ranks 1-13 and 14-26) sharing one scale for honest comparison.
from matplotlib.patches import Patch
def tight_col(v):
    return MAGENTA if v <= 25 else (A11Y if v <= 35 else BLUE)
allc = sorted(crows, key=lambda r: r['JOBSEEKERS'], reverse=True)
half = (len(allc) + 1) // 2          # 13 per column
groups = [allc[:half], allc[half:]]
xmax = max(r['JOBSEEKERS'] for r in allc)
fig, axes = plt.subplots(1, 2, figsize=(9.7, 4.15))
fig.subplots_adjust(wspace=0.55)
for ax, grp in zip(axes, groups):
    g = list(reversed(grp))           # highest at top
    names = [r['county'] for r in g]
    js    = [r['JOBSEEKERS'] for r in g]
    jspj  = [r['jobseekers_per_job'] for r in g]
    cols  = [tight_col(v) for v in jspj]
    ax.barh(range(len(g)), js, color=cols, height=0.66, zorder=2)
    ax.set_yticks(range(len(g)))
    ax.set_yticklabels(names, fontsize=9.2)
    style_ax(ax)
    ax.set_xlim(0, xmax * 1.02)
    ax.grid(axis='x', color=LGREY, linewidth=0.8)
    ax.tick_params(axis='x', labelsize=8.5)
    ax.xaxis.set_major_formatter(FuncFormatter(lambda v, p: f"{int(v/1000)}k" if v else "0"))
    for i, (v, jp) in enumerate(zip(js, jspj)):
        ax.text(v + xmax*0.015, i, f"{v:,.0f} \u00b7 {jp:.0f}/job", va='center',
                fontsize=7.6, color='#4A5160')
    ax.set_ylim(-0.6, len(g)-0.4)
axes[0].set_xlabel('Active jobseekers (care roles) \u00b7 seekers per job', fontsize=9.5)
axes[1].set_xlabel('Active jobseekers (care roles) \u00b7 seekers per job', fontsize=9.5)
leg = [Patch(color=MAGENTA, label='Tight supply (under 25 seekers/job)'),
       Patch(color=A11Y, label='Moderate (26\u201335 seekers/job)'),
       Patch(color=BLUE, label='Ample (over 35 seekers/job)')]
fig.legend(handles=leg, frameon=False, fontsize=9, loc='upper center',
           ncol=3, bbox_to_anchor=(0.5, 1.06))
save(fig, 'F_county_demand.png')

# ---------------------------------------------------------------- Chart G
# Maternity top-up benefit gap
labels_g = ['Irish private\nsector employers', 'Hospitality &\naccommodation', 'Private care\n(comparable)', 'Virtue']
vals_g   = [52, 20, 20, 100]
cols_g   = [GREY, GREY, GREY, BLUE]
fig, ax = plt.subplots(figsize=(8.0, 4.2))
b = ax.bar(range(len(labels_g)), vals_g, color=cols_g, width=0.62, zorder=2)
style_ax(ax)
ax.set_xticks(range(len(labels_g))); ax.set_xticklabels(labels_g, fontsize=9.3)
ax.get_xticklabels()[-1].set_color(BLUE); ax.get_xticklabels()[-1].set_weight('bold')
ax.set_ylabel('Offer maternity top-up (%)', fontsize=10.5)
ax.set_ylim(0, 116)
ax.grid(axis='y', color=LGREY, linewidth=0.9)
for i, v in enumerate(vals_g):
    txt = f"{v}%" + ("\u2713" if i==3 else "")
    ax.text(i, v + 2, ("Yes" if i==3 else f"~{v}%"), ha='center',
            fontsize=10, color=(BLUE if i==3 else DARK), weight=('bold' if i==3 else 'normal'))
save(fig, 'G_benefit_gap.png')

# ---------------------------------------------------------------- Chart H
# Marginal gains funnel: 5% relative improvement at each stage
stages = ['Job ad\nCTR', 'Apply-start\nrate', 'Application\ncompletion',
          'Interview\nattendance', 'Offer\nacceptance']
base = [3.0, 40, 60, 70, 75]
new  = [3.15, 42, 63, 73.5, 78.75]
fig, ax = plt.subplots(figsize=(8.4, 4.0))
xi = np.arange(len(stages)); w = 0.38
ax.bar(xi - w/2, base, w, label='Current benchmark', color=GREY, zorder=2)
ax.bar(xi + w/2, new,  w, label='+5% relative (benefit-led)', color=BLUE, zorder=2)
style_ax(ax)
ax.set_xticks(xi); ax.set_xticklabels(stages, fontsize=9.2)
ax.set_ylabel('Rate (%)', fontsize=10.5)
ax.grid(axis='y', color=LGREY, linewidth=0.9)
ax.legend(frameon=False, fontsize=9.6, loc='upper left')
ax.set_ylim(0, 92)
for xx, v in zip(xi + w/2, new):
    ax.text(xx, v + 1.5, f"{v:g}", ha='center', fontsize=8.6, color=BLUE, weight='bold')
save(fig, 'H_marginal_gains.png')

print('\nALL CHARTS BUILT ->', OUT)
print('market rate %.2f, virtue rate %.2f' % (mkt_rate, num(virt['Applystart_rate'])*100))
